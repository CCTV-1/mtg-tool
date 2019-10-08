import json
from enum import Enum

import requests
from openpyxl import Workbook

columnnames = ["编号", "中文名", "英语名", "类别", "法术力费用",
               "异能", "稀有度", "攻防", "现开", "轮抓", "构筑"]
raritytranslation = {"common": "普通",
                     "uncommon": "非普通", "rare": "稀有", "mythic": "秘稀"}


class Languages(Enum):
    EN = 'en'
    ES = 'es'
    FR = 'fr'
    DE = 'de'
    IT = 'it'
    PT = 'pt'
    JS = 'ja'
    KO = 'ko'
    RU = 'ru'
    ZHS = 'zhs'
    ZHT = 'zht'


class Card:
    def __init__(self, id: int, zh: str, en: str, mc: str, type: str, text: str, rarity: str, pt_value: str):
        self.set_id = id
        self.zh_name = zh
        self.en_name = en
        self.manacost = mc
        self.type_name = type
        self.text = text
        self.rarity = raritytranslation[rarity]
        self.pt = pt_value
        self.sealed_rating = 0
        self.darft_rating = 0
        self.construct_rating = 0


def write_xlsx(setcode: str, cards: list):
    wb = Workbook()
    sheet = wb.create_sheet("rating", 0)
    for column in range(0, columnnames.__len__()):
        sheet.cell(1, column + 1, columnnames[column])
    for number in range(0, cards.__len__()):
        sheet.cell(number + 2, 1, cards[number].set_id)
        sheet.cell(number + 2, 2, cards[number].zh_name)
        sheet.cell(number + 2, 3, cards[number].en_name)
        sheet.cell(number + 2, 4, cards[number].manacost)
        sheet.cell(number + 2, 5, cards[number].type_name)
        sheet.cell(number + 2, 6, cards[number].text)
        sheet.cell(number + 2, 7, cards[number].rarity)
        sheet.cell(number + 2, 8, cards[number].pt)
        sheet.cell(number + 2, 9, cards[number].sealed_rating)
        sheet.cell(number + 2, 10, cards[number].darft_rating)
        sheet.cell(number + 2, 11, cards[number].construct_rating)

    wb.save('{0}.xlsx'.format(setcode))


def get_cards(setcode: str, lang=Languages.EN):
    cards = []
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'}
    resurl = 'https://api.scryfall.com/cards/search?q=s:{0}+lang:{1}+is:booster'.format(
        setcode, lang.value)
    try:
        while True:
            resp = requests.get(resurl, timeout=60, headers=headers)
            if resp.status_code != 200:
                return None
            info_content = resp.json()
            for cardinfo in info_content['data']:
                id = cardinfo['collector_number']

                if 'card_faces' not in cardinfo:
                    if 'printed_name' in cardinfo:
                        locale_name = cardinfo['printed_name']
                    else:
                        locale_name = cardinfo['name']
                else:
                    cardface = cardinfo['card_faces'][0]
                    locale_name = ""
                    if 'printed_name' in cardface:
                        locale_name += cardface['printed_name']
                    else:
                        locale_name += cardface['name']
                    locale_name += ' // '
                    cardface = cardinfo['card_faces'][1]
                    if 'printed_name' in cardface:
                        locale_name += cardface['printed_name']
                    else:
                        locale_name += cardface['name']

                en_name = cardinfo['name']

                if 'printed_type_line' in cardinfo:
                    type_str = cardinfo['printed_type_line']
                else:
                    type_str = cardinfo['type_line']

                if 'mana_cost' in cardinfo:
                    mana_cost_str = cardinfo['mana_cost']
                else:
                    mana_cost_str = None

                if 'printed_text' in cardinfo:
                    card_text = cardinfo['printed_text']
                elif 'oracle_text' in cardinfo:
                    card_text = cardinfo['oracle_text']
                else:
                    if 'card_faces' not in cardinfo:
                        card_text = ""
                    else:
                        cardface = cardinfo['card_faces'][0]
                        card_text = ""
                        if 'printed_text' in cardface:
                            card_text += cardface['printed_text']
                        elif 'oracle_text' in cardface:
                            card_text += cardface['oracle_text']
                        cardface = cardinfo['card_faces'][1]
                        if card_text is not "":
                            card_text += '\n\n\n\n'
                        if 'printed_text' in cardface:
                            card_text += cardface['printed_text']
                        elif 'oracle_text' in cardface:
                            card_text += cardface['oracle_text']

                if 'power' in cardinfo:
                    card_pt = "{0}/{1}".format(cardinfo['power'],
                                               cardinfo['toughness'])
                else:
                    if 'card_faces' not in cardinfo:
                        card_pt = None
                    else:
                        cardface = cardinfo['card_faces'][0]
                        card_pt = ""
                        if 'power' in cardface:
                            card_pt += "{0}/{1}".format(cardface['power'],
                                                        cardface['toughness'])
                            card_pt += '\n'
                        cardface = cardinfo['card_faces'][1]
                        if 'power' in cardface:
                            card_pt += "{0}/{1}".format(cardface['power'],
                                                        cardface['toughness'])

                cards.append(Card(id, locale_name, en_name, type_str, mana_cost_str,
                                  card_text, cardinfo['rarity'], card_pt))
            if 'has_more' in info_content:
                has_more = info_content['has_more']
                if has_more == True:
                    resurl = info_content['next_page']
                else:
                    break
            else:
                break

        return cards
    except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectTimeout):
        print(cardinfo)
        return cards
    except (AttributeError, TypeError, KeyError):
        print(cardinfo)
        return cards


if __name__ == "__main__":
    code = input()
    cards = get_cards(code, Languages.ZHS)
    write_xlsx(code, cards)
