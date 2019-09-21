import json
import logging

import requests
from openpyxl import Workbook

columnnames = ["中文名", "英语名", "类别", "法术力费用",
               "异能", "稀有度", "攻防", "现开", "轮抓", "构筑"]
raritytranslation = {"common": "普通",
                     "uncommon": "非普通", "rare": "稀有", "mythic": "秘稀"}


class Card:
    def __init__(self, zh: str, en: str, mc: str, type: str, text: str, rarity: str, pt_value: str):
        self.zh_name = zh
        self.en_name = en
        self.manacost = mc
        self.type_name = type
        self.text = text
        self.rarity = raritytranslation[rarity]
        self.pt = pt_value
        self.sealed_ranking = 0
        self.darft_ranking = 0
        self.construct_ranking = 0


def write_xlsx(setcode: str, cards: list):
    wb = Workbook()
    sheet = wb.create_sheet("ranking", 0)
    for column in range(0, columnnames.__len__()):
        sheet.cell(1, column + 1, columnnames[column])
    for number in range(0, cards.__len__()):
        sheet.cell(number + 2, 1, cards[number].zh_name)
        sheet.cell(number + 2, 2, cards[number].en_name)
        sheet.cell(number + 2, 3, cards[number].manacost)
        sheet.cell(number + 2, 4, cards[number].type_name)
        sheet.cell(number + 2, 5, cards[number].text)
        sheet.cell(number + 2, 6, cards[number].rarity)
        sheet.cell(number + 2, 7, cards[number].pt)
        sheet.cell(number + 2, 8, cards[number].sealed_ranking)
        sheet.cell(number + 2, 9, cards[number].darft_ranking)
        sheet.cell(number + 2, 10, cards[number].construct_ranking)

    wb.save('{0}.xlsx'.format(setcode))


def get_cards(setcode: str, lang='en'):
    cards = []
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'}
    try:
        resp = requests.get(
            'https://api.scryfall.com/cards/search?q=s:{0}+lang:{1}'.format(setcode, lang), timeout=60, headers=headers)
        if resp.status_code != 200:
            return None
        info_content = resp.json()
        for cardinfo in info_content['data']:
            if 'printed_name' in cardinfo:
                locale_name = cardinfo['printed_name']
            else:
                locale_name = cardinfo['name']

            en_name = cardinfo['name']

            if 'printed_type_line' in cardinfo:
                type_str = cardinfo['printed_type_line']
            else:
                type_str = cardinfo['type_line']

            if 'mana_cost' in cardinfo:
                mana_cost_str = cardinfo['mana_cost']
            else:
                mana_cost_str = None

            if 'printed_text' in cardinfo:
                card_text = cardinfo['printed_text']
            elif 'oracle_text' in cardinfo:
                card_text = cardinfo['oracle_text']
            else:
                card_text = None

            if 'power' in cardinfo:
                card_pt = "{0}/{1}".format(cardinfo['power'],
                                           cardinfo['toughness'])
            else:
                card_pt = None

            cards.append(Card(locale_name, en_name, type_str, mana_cost_str,
                              card_text, cardinfo['rarity'], card_pt))
        has_more = info_content['has_more']

        while has_more != False:
            resp = requests.get(
                info_content.get('next_page'), timeout=60, headers=headers)
            info_content = resp.json()
            for cardinfo in info_content['data']:
                if 'printed_name' in cardinfo:
                    locale_name = cardinfo['printed_name']
                else:
                    locale_name = cardinfo['name']

                en_name = cardinfo['name']

                if 'printed_type_line' in cardinfo:
                    type_str = cardinfo['printed_type_line']
                else:
                    type_str = cardinfo['type_line']

                if 'mana_cost' in cardinfo:
                    mana_cost_str = cardinfo['mana_cost']
                else:
                    mana_cost_str = None

                if 'printed_text' in cardinfo:
                    card_text = cardinfo['printed_text']
                elif 'oracle_text' in cardinfo:
                    card_text = cardinfo['oracle_text']
                else:
                    card_text = None

                if 'power' in cardinfo:
                    card_pt = "{0}/{1}".format(cardinfo['power'],
                                               cardinfo['toughness'])
                else:
                    card_pt = None

                cards.append(Card(locale_name, en_name, type_str, mana_cost_str,
                                  card_text, cardinfo['rarity'], card_pt))
            has_more = info_content['has_more']

        return cards
    except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectTimeout):
        logging.info("%s", str(cardinfo))
        return cards
    except (AttributeError, TypeError, KeyError):
        logging.info("%s", str(cardinfo))
        return cards


if __name__ == "__main__":
    logging.basicConfig(filename='generator.log',
                        filemode='w', level=logging.INFO)
    code = input()
    cards = get_cards(code, 'en')
    write_xlsx(code, cards)
